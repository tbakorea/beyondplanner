from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

ROOT = Path("/Users/bangju/Documents/비욘드 플래너")
XLSX = ROOT / "outputs/franklin_planner_final/2026_ㅇㅇㅇ_프랭클린플래너_개선판.xlsx"


def clear_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def set_row(ws, row_index, values):
    for col, value in enumerate(values, 1):
        ws.cell(row_index, col, value)


wb = openpyxl.load_workbook(XLSX)

if "업무관리" in wb.sheetnames:
    ws = wb["업무관리"]
    clear_range(ws, 7, 506, 1, 12)
    ws["A2"] = "업무 원본 입력 공간입니다. 개인정보와 업무일정은 제거되었습니다."

if "시간관리" in wb.sheetnames:
    ws = wb["시간관리"]
    clear_range(ws, 7, 756, 1, 7)
    ws["A2"] = "시간표 원본 입력 공간입니다. 일정 정보는 제거되었습니다."

if "오늘" in wb.sheetnames:
    ws = wb["오늘"]
    ws["A2"] = "개인정보 제거 후의 빈 대시보드 화면입니다."
    ws["B4"] = None
    ws["D4"] = None
    ws["B5"] = 0
    ws["E5"] = 0
    clear_range(ws, 9, 48, 1, 7)
    clear_range(ws, 54, 82, 1, 5)

if "월간반복" in wb.sheetnames:
    ws = wb["월간반복"]
    ws["A2"] = "반복 일정과 민감 정보는 제거되었습니다."
    clear_range(ws, 7, 80, 1, 6)

if "패턴분석" in wb.sheetnames:
    ws = wb["패턴분석"]
    clear_range(ws, 6, 30, 1, 4)
    ws._charts = []
    set_row(ws, 6, ["개인정보 제거", "완료", "개인정보, 업무일정 및 민감 정보는 산출물에서 제거되었습니다.", "새 데이터를 직접 입력하면 대시보드가 다시 집계됩니다."])
    set_row(ws, 7, ["업무 데이터", "비어 있음", "사용자 업무 내용은 표시하지 않습니다.", "업무관리 시트에 직접 입력하세요."])
    set_row(ws, 8, ["시간표 데이터", "비어 있음", "사용자 시간 일정은 표시하지 않습니다.", "시간관리 시트에 직접 입력하세요."])

if "수식맵" in wb.sheetnames:
    ws = wb["수식맵"]
    clear_range(ws, 6, 10, 1, 4)
    rows = [
        ["연간/월간 날짜", "=이전 날짜+1", "날짜 셀을 연속 증가시켜 연간, 월간 캘린더를 만듭니다.", "개인 데이터 없이 구조만 유지합니다."],
        ["업무 완료 처리", "조건부서식", "완료/취소/연기/위임 상태에서 취소선을 적용합니다.", "업무관리 시트의 직접 입력값에만 적용됩니다."],
        ["미완료 이월", "INDEX + AGGREGATE", "해당일 또는 이전 미완료/진행중/연기 업무를 오늘 화면으로 끌어옵니다.", "참고 사용자 업무는 제거되었습니다."],
        ["시간별 일정", "INDEX + SUMPRODUCT", "선택일과 시간 슬롯이 일치하는 일정만 표시합니다.", "참고 사용자 일정은 제거되었습니다."],
        ["패턴 분석", "COUNTIFS / SUM", "직접 입력한 데이터로만 집계합니다.", "민감 정보는 포함하지 않습니다."],
    ]
    for offset, row in enumerate(rows, 6):
        set_row(ws, offset, row)

if "출처" in wb.sheetnames:
    ws = wb["출처"]
    clear_range(ws, 6, 8, 1, 2)
    set_row(ws, 6, ["공개용 프랭클린플래너 구조", "연간 일정, 목표설정, 월별 주간/일간 플래너 구조 참고"])
    set_row(ws, 7, ["익명화 처리", "개인정보, 업무일정 및 민감 정보 제거"])
    set_row(ws, 8, ["개선판 설계", "업무관리와 시간관리를 원본 테이블로 두고 오늘 화면이 자동 집계되는 구조"])

for sheet_name in wb.sheetnames:
    if sheet_name.endswith("월"):
        ws = wb[sheet_name]
        for row in range(6, ws.max_row + 1, 2):
            for col in range(1, 8):
                ws.cell(row, col).value = None

wb.save(XLSX)
print(XLSX)
