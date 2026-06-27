from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path("/Users/bangju/Documents/프랭클린 플래너 생성")
SOURCE = ROOT / "outputs/franklin_planner_final/2026_ㅇㅇㅇ_프랭클린플래너_개선판.xlsx"
OUT = ROOT / "dashboard/data.js"


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def read_table(ws, header_row: int, start_row: int, end_row: int, start_col: int, end_col: int):
    headers = [clean(ws.cell(header_row, col).value) for col in range(start_col, end_col + 1)]
    rows = []
    for r in range(start_row, end_row + 1):
        row = {headers[i]: clean(ws.cell(r, start_col + i).value) for i in range(len(headers))}
        if any(v not in ("", None) for v in row.values()):
            rows.append(row)
    return rows


def as_number(value: Any) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return 0.0


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, data_only=True)

    tasks = read_table(wb["업무관리"], 6, 7, 506, 1, 12)
    schedule = read_table(wb["시간관리"], 6, 7, 756, 1, 7)
    repeats = read_table(wb["월간반복"], 6, 7, 80, 1, 6)
    pattern = read_table(wb["패턴분석"], 5, 6, 11, 1, 4)

    today_ws = wb["오늘"]
    today = clean(today_ws["B4"].value)
    today_tasks = read_table(today_ws, 8, 9, 48, 1, 7)
    today_schedule = read_table(today_ws, 53, 54, 82, 1, 5)

    status_counts: dict[str, int] = {}
    priority_counts: dict[str, int] = {}
    role_counts: dict[str, int] = {}
    carry_count = 0
    for task in tasks:
        if not task.get("업무내용"):
            continue
        status = str(task.get("상태") or "미완료")
        priority = str(task.get("우선순위") or "-")
        role = str(task.get("역할") or "미분류")
        status_counts[status] = status_counts.get(status, 0) + 1
        priority_counts[priority] = priority_counts.get(priority, 0) + 1
        role_counts[role] = role_counts.get(role, 0) + 1
        if task.get("이월여부") == "이월":
            carry_count += 1

    filled_slots = [row for row in schedule if row.get("일정내용")]
    completed = status_counts.get("완료", 0)
    active_total = sum(status_counts.values())
    blocked_like = sum(status_counts.get(s, 0) for s in ["미완료", "진행중", "연기"])
    completion_rate = completed / active_total if active_total else 0

    repeat_days: dict[str, int] = {}
    for row in repeats:
        day = str(row.get("일") or "-")
        repeat_days[day] = repeat_days.get(day, 0) + 1

    metrics = {
        "today": today,
        "taskTotal": active_total,
        "completionRate": completion_rate,
        "carryCount": carry_count,
        "openCount": blocked_like,
        "filledSlots": len(filled_slots),
        "repeatCount": len(repeats),
    }

    payload = {
        "source": {
            "file": str(SOURCE),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "freshness": "개선판 엑셀에서 추출",
        },
        "metrics": metrics,
        "tasks": tasks,
        "todayTasks": today_tasks,
        "schedule": schedule,
        "todaySchedule": today_schedule,
        "repeats": [
            {key: value for key, value in row.items() if key != "금액"}
            for row in repeats
        ],
        "pattern": pattern,
        "charts": {
            "statusCounts": status_counts,
            "priorityCounts": priority_counts,
            "roleCounts": role_counts,
            "repeatDays": repeat_days,
        },
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "window.FRANKLIN_DASHBOARD_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(OUT)


if __name__ == "__main__":
    main()
